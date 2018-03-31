from openpyxl import load_workbook
from openpyxl import Workbook
from slugify import slugify
from os.path import basename

class xlsx2tab:
    def __init__ (self):
        pass
    def output_tab(self, options):
        wb = load_workbook(options.input) 
        if not options.prefix:
            options.prefix = basename(options.input)
        sheetnames = wb.get_sheet_names()
        for sheetname in sheetnames:
            sn_slug = slugify(sheetname)            
            output_file = '{}.{}.{}'.format(options.prefix, sn_slug, "txt")
            with open(output, "w") as output_file:
                sheet = wb[sheetname]
                for row in sheet.iter_rows():
                    ele = []
                    for cell in row:
                        ele.append(str(cell.value))
                    output_file.write("\t".join(ele))
