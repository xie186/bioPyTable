from openpyxl import load_workbook
from openpyxl import Workbook
from slugify import slugify
from os.path import basename

class tab2xlsx:
    def __init__ (self):
        pass

    def output_xlsx(self, options):
        wb = Workbook()
        tab_list = options.tab.split(",")
        sheet_list = tab_list
        if options.sheet:
            sheet_list = options.sheet.split(",")
        if len(tab_list) != len(sheet_list):
            raise Exception("Sheets number is not equal to table number! Please check!")
        for i in range(0, len(sheet_list)):
            ws = wb.create_sheet(sheet_list[i], i)
            ws.freeze_panes = options.freeze_panes
            with open(tab_list[i]) as tab:
                row_index = 1
                for line in tab:
                    line = line.rstrip()
                    ele = line.split("\t")
                    for j in range(0, len(ele)):
                        #print(row_index)
                        ws.cell(row=row_index, column=j+1).value = ele[j]
                    row_index += 1
        wb.save(options.output)
                    
